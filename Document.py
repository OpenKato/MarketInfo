# coding: utf-8
import sys
#import openpyxl, pprint
import openpyxl as px
from datetime import datetime

# ファイル名の指定など --- (*1)
file_master = "file_master.xlsx" # マスターデータ
touhoku = "touhoku.xlsx" # 東北のデータ
kita = "kita.xlsx"       # 北関東のデータ
minami = "minami.xlsx"   # 南関東のデータ
shizuoka  = "shizuoka.xlsx" # 静岡のデータ
hiroshima = "hiroshima.xlsx" # 広島のデータ
fukuoka ="fukuoka.xlsx"      # 福岡のデータ
tokyo ="tokyo.xlsx"      # 東京のデータ
nagoya ="nagoya.xlsx"      # 名古屋のデータ
osaka ="osaka.xlsx"      # 大阪のデータ

file_master82 = "file_master82-ver2.xlsx"

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master2)
#print("ok")

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master2.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master3)
print("ok")

# 北関東データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master3.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=201+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master4)
print("ok")

# 北関東データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master4.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=201+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master5)
print("ok")

# 南関東データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master5.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=401+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master6)
print("ok")

# 南関東データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master6.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=401+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master7)
print("ok")

# 静岡データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master7.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=601+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master8)
print("ok")

# 静岡データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master8.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=601+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master9)
print("ok")

# 広島データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master9.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=801+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master10)
print("ok")

# 広島データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master10.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=801+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master11)
print("ok")

# 福岡データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master11.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1001+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master12)
print("ok")

# 福岡データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master12.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1001+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master13)
print("ok")

# 東京データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master13.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1201+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master14)
print("ok")

# 東京データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master14.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1201+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master15)
print("ok")

# 名古屋データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master15.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1401+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master16)
print("ok")

# 名古屋データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master16.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1401+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master17)
print("ok")

# 大阪データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["May"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master17.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1601+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master18)
print("ok")

# 大阪データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Jun"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master18.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1601+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master19)
print("ok")

#当月受注、来月商況は終了！！！！！！！！！！！
#ここからは、現場ニュース
#東北

#失注情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A3:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master19.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master20)
print("ok")

#競合情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master20.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master21)
print("ok")

#商社情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master21.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master22)
print("ok")

#顧客情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master22.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master23)
print("ok")

#他部門依頼情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A53:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master23.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master24)
print("ok")

#クレーム情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master24.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master25)
print("ok")

#クレーム情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A94:M104"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master25.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master26)
print("ok")

#北関東

#失注情報

# データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A4:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master26.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=11+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master27)
print("ok")

#競合情報

# データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master27.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=11+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master28)
print("ok")

#商社情報

# データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master28.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=11+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master29)
print("ok")

#顧客情報

# データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master29.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=11+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master30)
print("ok")

#他部門依頼情報

# データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A54:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master30.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=11+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master31)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master31.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=11+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master32)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A96:M106"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master32.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=11+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master33)
print("ok")

#東京

#失注情報

# データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A4:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master33.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=21+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master34)
print("ok")

#競合情報

# データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master34.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=21+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master35)
print("ok")

#商社情報

# データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master35.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=21+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master36)
print("ok")

#顧客情報

# データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master36.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=21+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master37)
print("ok")

#他部門依頼情報

# データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A54:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master37.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=21+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master38)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master38.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=21+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master39)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A82:M92"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('339.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=21+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master40)
print("ok")

#南関東

#失注情報

# データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A4:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master40.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=31+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master41)
print("ok")

#競合情報

# データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M29"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master41.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=31+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master42)
print("ok")

#商社情報

# データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A33:M57"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master42.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=31+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master43)
print("ok")

#顧客情報

# データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A61:M85"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master43.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=31+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master44)
print("ok")

#他部門依頼情報

# データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A89:M98"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master44.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=31+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master45)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A102:M111"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master45.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=31+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master46)
print("ok")

#海外支援

# データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A131:M141"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master46.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=31+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master47)
print("ok")

#静岡

#失注情報

# データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A4:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master47.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=41+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master48)
print("ok")

#競合情報

# データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master48.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=46+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master49)
print("ok")

#商社情報

# データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master49.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=56+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master50)
print("ok")

#顧客情報

# データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master50.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=56+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master51)
print("ok")

#他部門依頼情報

# データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A54:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master51.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=41+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master52)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master52.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=41+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master53)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A96:M106"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master53.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=41+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master54)
print("ok")

#名古屋

#失注情報

# データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A4:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('54.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=51+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master55)
print("ok")

#競合情報

# データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master55.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=56+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master56)
print("ok")

#商社情報

# データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master56.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=66+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master57)
print("ok")

#顧客情報

# データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master57.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=66+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master58)
print("ok")

#他部門依頼情報

# データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A54:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master58.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=51+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master59)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master59.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=51+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master60)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A82:M92"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master60.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=51+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master61)
print("ok")


#大阪

#失注情報

# データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A4:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master61.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=61+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master62)
print("ok")

#競合情報

# データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master62.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=66+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master63)
print("ok")

#商社情報

# データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master63.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=76+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master64)
print("ok")

#顧客情報

# データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master64.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=76+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master65)
print("ok")

#他部門依頼情報

# データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A54:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master65.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=61+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master66)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master66.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=61+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master67)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A82:M92"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master67.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=51+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master68)
print("ok")

#広島

#失注情報

# データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A4:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master68.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=71+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master69)
print("ok")

#競合情報

# データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master69.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=76+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master70)
print("ok")

#商社情報

# データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master70.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=86+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master71)
print("ok")

#顧客情報

# データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master71.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=86+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master72)
print("ok")

#他部門依頼情報

# データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A54:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master72.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=71+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master73)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master73.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=71+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master74)
print("ok")

#海外情報

# データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A96:M106"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master74.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=61+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master75)
print("ok")

#福岡

#失注情報

# データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A4:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master75.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=81+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master76)
print("ok")

#競合情報

# データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master76.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=86+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master77)
print("ok")

#商社情報

# データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master77.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=96+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master78)
print("ok")

#顧客情報

# データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master78.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=96+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master79)
print("ok")

#他部門依頼情報

# データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A54:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master79.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=81+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master80)
print("ok")

#クレーム情報

# データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master80.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=81+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
#wb_iv.save(file_master81)
print("ok")

#海外情報

# データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A96:M106"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
#wb_iv = px.load_workbook('file_master81.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=71+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_master82)
print("last-OK")
