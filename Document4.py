# -*- coding: utf-8 -*-
#import openpyxl, pprint
import openpyxl as px
from datetime import datetime

# ファイル名の指定など --- (*1)
file_master = "file_master.xlsx" # マスターデータ
touhoku = "touhoku.xlsx" # 東北のデータ
kita = "kita.xlsx"       # 北関東のデータ
minami = "minami.xlsx"   # 南関東のデータ
shizuoka  = "shizuoka.xlsx" # 静岡のデータ
hiroshima = "hiroshima.xlsx" # 広島のデータ
fukuoka ="fukuoka.xlsx"      # 福岡のデータ
tokyo ="tokyo.xlsx"      # 東京のデータ
nagoya ="nagoya.xlsx"      # 名古屋のデータ
osaka ="osaka.xlsx"      # 大阪のデータ


file_master2 = "file_master2.xlsx"
file_master3 = "file_master3.xlsx"
file_master4 = "file_master4.xlsx"
file_master5 = "file_master5.xlsx"
file_master6 = "file_master6.xlsx"
file_master7 = "file_master7.xlsx"
file_master8 = "file_master8.xlsx"
file_master9 = "file_master9.xlsx"
file_master10 = "file_master10.xlsx"
file_master11 = "file_master11.xlsx"
file_master12 = "file_master12.xlsx"
file_master13 = "file_master13.xlsx"
file_master14 = "file_master14.xlsx"
file_master15 = "file_master15.xlsx"
file_master16 = "file_master16.xlsx"
file_master17 = "file_master17.xlsx"
file_master18 = "file_master18.xlsx"
file_master19 = "file_master19.xlsx"
file_master20 = "file_master20.xlsx"
file_master21 = "file_master21.xlsx"
file_master22 = "file_master22.xlsx"
file_master23 = "file_master23.xlsx"
file_master24 = "file_master24.xlsx"
file_master25 = "file_master25.xlsx"
file_master26 = "file_master26.xlsx"
file_master27 = "file_master27.xlsx"
file_master28 = "file_master28.xlsx"
file_master29 = "file_master29.xlsx"
file_master30 = "file_master30.xlsx"
file_master31 = "file_master31.xlsx"
file_master32 = "file_master32.xlsx"
file_master33 = "file_master33.xlsx"
file_master34 = "file_master34.xlsx"
file_master35 = "file_master35.xlsx"
file_master36 = "file_master36.xlsx"
file_master37 = "file_master37.xlsx"
file_master38 = "file_master38.xlsx"
file_master39 = "file_master39.xlsx"
file_master40 = "file_master40.xlsx"
file_master41 = "file_master41.xlsx"
file_master42 = "file_master42.xlsx"
file_master43 = "file_master43.xlsx"
file_master44 = "file_master44.xlsx"
file_master45 = "file_master45.xlsx"
file_master46 = "file_master46.xlsx"
file_master47 = "file_master47.xlsx"
file_master48 = "file_master48.xlsx"
file_master49 = "file_master49.xlsx"
file_master50 = "file_master50.xlsx"
file_master51 = "file_master51.xlsx"


# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master2)
print("ok")

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master2.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master3)
print("ok")

# 北関東データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master3.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=201+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master4)
print("ok")

# 北関東データを読み込む --- (*2)
wb = px.load_workbook(kita, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master4.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=201+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master5)
print("ok")

# 南関東データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master5.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=401+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master6)
print("ok")

# 南関東データを読み込む --- (*2)
wb = px.load_workbook(minami, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master6.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=401+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master7)
print("ok")

# 静岡データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master7.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=601+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master8)
print("ok")

# 静岡データを読み込む --- (*2)
wb = px.load_workbook(shizuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master8.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=601+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master9)
print("ok")

# 広島データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master9.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=801+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master10)
print("ok")

# 広島データを読み込む --- (*2)
wb = px.load_workbook(hiroshima, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master10.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=801+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master11)
print("ok")

# 福岡データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master11.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1001+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master12)
print("ok")

# 福岡データを読み込む --- (*2)
wb = px.load_workbook(fukuoka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master12.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1001+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master13)
print("ok")

# 東京データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master13.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1201+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master14)
print("ok")

# 東京データを読み込む --- (*2)
wb = px.load_workbook(tokyo, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master14.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1201+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master15)
print("ok")

# 名古屋データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master15.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1401+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master16)
print("ok")

# 名古屋データを読み込む --- (*2)
wb = px.load_workbook(nagoya, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master16.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1401+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master17)
print("ok")

# 大阪データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Nov"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master17.xlsx')
ws_iv = wb_iv["Sheet9"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1601+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master18)
print("ok")

# 大阪データを読み込む --- (*2)
wb = px.load_workbook(osaka, data_only=True) # 数式でなく値を取り出す場合
ws = wb["Dec"] # シート名を選ぶ
list_data = ws["A14:AX200"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master18.xlsx')
ws_iv = wb_iv["Sheet10"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1601+y+1, column=0+x+1, value=v)
   
# 新しく保存する --- (*6)
wb_iv.save(file_master19)
print("ok")

#当月受注、来月商況は終了！！！！！！！！！！！
#ここからは、現場ニュース
#東北

#失注情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A3:M11"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master19.xlsx')
ws_iv = wb_iv["shi"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_master20)
print("ok")

#競合情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A15:M24"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master20.xlsx')
ws_iv = wb_iv["kyo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_master21)
print("ok")

#商社情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A28:M37"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master21.xlsx')
ws_iv = wb_iv["syo"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_master22)
print("ok")

#顧客情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A41:M50"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master22.xlsx')
ws_iv = wb_iv["ko"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_master23)
print("ok")

#他部門依頼情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A54:M63"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master23.xlsx')
ws_iv = wb_iv["ta"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_master24)
print("ok")

#クレーム情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A67:M76"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master24.xlsx')
ws_iv = wb_iv["ku"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_master25)
print("ok")

#クレーム情報

# 東北データを読み込む --- (*2)
wb = px.load_workbook(touhoku, data_only=True) # 数式でなく値を取り出す場合
ws = wb["News"] # シート名を選ぶ
list_data = ws["A99:M106"] # 任意の範囲を取得

# マスタデータを読む --- (*3)
wb_iv = px.load_workbook('file_master25.xlsx')
ws_iv = wb_iv["ka"]

# 納品物を書き込む --- (*5)
for y, row in enumerate(list_data):
  for x, cell in enumerate(row):
    if (cell is None) or (cell.value is None): continue
    v = cell.value
    ws_iv.cell(row=1+y+1, column=0+x+1, value=v)

# 新しく保存する --- (*6)
wb_iv.save(file_master26)
print("ok")







